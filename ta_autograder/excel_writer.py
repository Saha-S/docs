"""
excel_writer.py – Write grading results to an Excel file.

Column layout (exactly as specified):
  Etunimi | Sukunimi | Tunnistenumero | Sähköpostiosoite |
  Task2 total (15p) | Tool points |
  Tool1 | Tool2 | Tool3 | Tool4 | Tool5 | Tool6 |
  Bonus/minus(1p) | (empty) | Comment
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

# Column headers in order
HEADERS = [
    "Etunimi",
    "Sukunimi",
    "Tunnistenumero",
    "Sähköpostiosoite",
    "Task2 total (15p)",
    "Tool points",
    "Tool1",
    "Tool2",
    "Tool3",
    "Tool4",
    "Tool5",
    "Tool6",
    "Bonus/minus(1p)",
    "",          # intentionally empty column
    "Comment",
]


def write_results(rows: list[dict[str, Any]], output_path: Path) -> None:
    """Write *rows* to an Excel file at *output_path*.

    Each dict in *rows* must have the keys produced by ``main.py``:
      first_name, last_name, student_id, email,
      final_score, tool_points,
      tool_scores (list of up to 6 floats),
      bonus_minus, comment
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grading"

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for row_idx, data in enumerate(rows, start=2):
        tool_scores: list[float | None] = data.get("tool_scores", [])
        # Pad to 6 entries
        padded_tools: list[float | None] = list(tool_scores) + [None] * (6 - len(tool_scores))
        padded_tools = padded_tools[:6]

        row_values: list[Any] = [
            data.get("first_name", ""),
            data.get("last_name", ""),
            data.get("student_id", ""),
            data.get("email", ""),
            data.get("final_score", ""),
            data.get("tool_points", ""),
            *padded_tools,
            data.get("bonus_minus", ""),
            "",  # empty column
            data.get("comment", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ------------------------------------------------------------------
    # Column widths (approximate)
    # ------------------------------------------------------------------
    col_widths = [12, 15, 16, 28, 18, 12,
                  8, 8, 8, 8, 8, 8,
                  16, 4, 60]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Results saved to: {output_path}")
